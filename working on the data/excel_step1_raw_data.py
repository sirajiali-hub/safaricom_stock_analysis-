import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

df = pd.read_csv('safaricom_cleaned.csv')
df['Date'] = pd.to_datetime(df['Date'])

wb = Workbook()
sheet = wb.active
sheet.title = 'Raw Data'

sheet['A1'] = 'Safaricom Stock Data - Raw Dataset'
sheet['A1'].font = Font(bold=True, size=14)

headers = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')

for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=3, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for idx, row in df[headers].iterrows():
    for col, value in enumerate(row, 1):
        sheet.cell(row=idx+4, column=col, value=value)

sheet.column_dimensions['A'].width = 12
for col in ['B', 'C', 'D', 'E']:
    sheet.column_dimensions[col].width = 12
sheet.column_dimensions['F'].width = 15

wb.save('safaricom_excel.xlsx')
