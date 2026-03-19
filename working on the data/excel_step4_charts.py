from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font

wb = load_workbook('safaricom_excel.xlsx')
charts_sheet = wb.create_sheet('Charts')

charts_sheet['A1'] = 'Visual Analysis'
charts_sheet['A1'].font = Font(bold=True, size=14)

analysis_sheet = wb['Analysis']

chart1 = LineChart()
chart1.title = "Stock Price Trend"
chart1.style = 10
chart1.y_axis.title = "Price (KES)"
chart1.x_axis.title = "Date"
chart1.height = 10
chart1.width = 20

data = Reference(analysis_sheet, min_col=2, min_row=3, max_row=526)
dates = Reference(analysis_sheet, min_col=1, min_row=4, max_row=526)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(dates)

charts_sheet.add_chart(chart1, "B3")

chart2 = LineChart()
chart2.title = "Moving Averages Comparison"
chart2.style = 12
chart2.y_axis.title = "Price (KES)"
chart2.x_axis.title = "Date"
chart2.height = 10
chart2.width = 20

data2 = Reference(analysis_sheet, min_col=2, min_row=3, max_col=5, max_row=526)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(dates)

charts_sheet.add_chart(chart2, "B25")

wb.save('safaricom_excel.xlsx')