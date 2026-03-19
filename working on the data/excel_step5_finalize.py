from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook('safaricom_excel.xlsx')

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

dashboard = wb['Dashboard']
for row in range(7, 17):
    for col in ['B', 'D']:
        dashboard[f'{col}{row}'].border = border

analysis = wb['Analysis']
for row in range(3, 527):
    for col in range(1, 7):
        cell = analysis.cell(row=row, column=col)
        cell.border = border

raw_data = wb['Raw Data']
for row in range(3, 527):
    for col in range(1, 7):
        cell = raw_data.cell(row=row, column=col)
        cell.border = border

info_sheet = wb.create_sheet('Info')
info_sheet['A1'] = 'Safaricom Stock Analysis - Excel Report'
info_sheet['A1'].font = Font(bold=True, size=16)
info_sheet['A3'] = 'Project: 2-Year Stock Market Analysis'
info_sheet['A4'] = 'Company: Safaricom PLC (NSE: SCOM)'
info_sheet['A5'] = 'Analysis Period: 2024-2026'
info_sheet['A7'] = 'Sheets:'
info_sheet['A8'] = '1. Dashboard - Key performance metrics'
info_sheet['A9'] = '2. Raw Data - Complete dataset'
info_sheet['A10'] = '3. Analysis - Calculated metrics'
info_sheet['A11'] = '4. Charts - Visual analysis'

info_sheet.column_dimensions['A'].width = 40

wb.save('safaricom_excel.xlsx')
print("Excel workbook complete!")