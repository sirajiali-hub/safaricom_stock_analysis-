from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

df = pd.read_csv('safaricom_cleaned.csv')
df['Date'] = pd.to_datetime(df['Date'])

wb = load_workbook('safaricom_excel.xlsx')
summary_sheet = wb.create_sheet('Dashboard', 0)

summary_sheet['B2'] = 'SAFARICOM STOCK PERFORMANCE DASHBOARD'
summary_sheet['B2'].font = Font(bold=True, size=16, color='1E3C72')
summary_sheet.merge_cells('B2:E2')

summary_sheet['B4'] = 'KEY METRICS'
summary_sheet['B4'].font = Font(bold=True, size=14)

input_font = Font(color='0000FF')
formula_font = Font(color='000000')
bold_font = Font(bold=True)

metrics = [
    ('Start Date', df['Date'].min().strftime('%Y-%m-%d')),
    ('End Date', df['Date'].max().strftime('%Y-%m-%d')),
    ('Start Price (KES)', df['Close'].iloc[0]),
    ('Current Price (KES)', df['Close'].iloc[-1]),
    ('Highest Price (KES)', df['Close'].max()),
    ('Lowest Price (KES)', df['Close'].min()),
    ('', ''),
    ('Total Return (%)', '=(D11-D10)/D10*100'),
    ('Trading Days', len(df)),
    ('Average Volume', df['Volume'].mean()),
]

for i, (label, value) in enumerate(metrics):
    row = 7 + i
    summary_sheet[f'B{row}'] = label
    summary_sheet[f'B{row}'].font = input_font
    
    cell = summary_sheet[f'D{row}']
    if isinstance(value, str) and value.startswith('='):
        cell.value = value
        cell.font = formula_font
    else:
        cell.value = value
        cell.font = bold_font

summary_sheet['D10'].number_format = '#,##0.00'
summary_sheet['D11'].number_format = '#,##0.00'
summary_sheet['D12'].number_format = '#,##0.00'
summary_sheet['D13'].number_format = '#,##0.00'
summary_sheet['D14'].number_format = '0.00%'
summary_sheet['D15'].number_format = '#,##0'
summary_sheet['D16'].number_format = '#,##0'

summary_sheet.column_dimensions['B'].width = 25
summary_sheet.column_dimensions['D'].width = 20

wb.save('safaricom_excel.xlsx')