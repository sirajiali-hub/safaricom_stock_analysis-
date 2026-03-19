import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

df = pd.read_csv('safaricom_cleaned.csv')
df['Date'] = pd.to_datetime(df['Date'])

wb = load_workbook('safaricom_excel.xlsx')
analysis_sheet = wb.create_sheet('Analysis')

analysis_sheet['A1'] = 'Safaricom Stock Analysis'
analysis_sheet['A1'].font = Font(bold=True, size=14)

headers = ['Date', 'Close Price', 'Daily Return (%)', 'MA 7-Day', 'MA 30-Day', 'Volatility 30D']
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')

for col, header in enumerate(headers, 1):
    cell = analysis_sheet.cell(row=3, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for idx, row in df[['Date', 'Close', 'Daily_Return', 'MA_7', 'MA_30', 'Volatility_30']].iterrows():
    analysis_sheet.cell(row=idx+4, column=1, value=row['Date'])
    analysis_sheet.cell(row=idx+4, column=2, value=row['Close'])
    analysis_sheet.cell(row=idx+4, column=3, value=row['Daily_Return'])
    analysis_sheet.cell(row=idx+4, column=4, value=row['MA_7'])
    analysis_sheet.cell(row=idx+4, column=5, value=row['MA_30'])
    analysis_sheet.cell(row=idx+4, column=6, value=row['Volatility_30'])

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    analysis_sheet.column_dimensions[col].width = 15

wb.save('safaricom_excel.xlsx')